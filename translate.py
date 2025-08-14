import openpyxl
import os
import sys
import json  # 追加
import re  # 日本語判定用
from openai import AzureOpenAI
from dotenv import load_dotenv

# Azure OpenAI API key and endpoint
# Load environment variables from .env file
load_dotenv()

api_key = os.getenv("AZURE_OPENAI_API_KEY")
endpoint = os.getenv("AZURE_OPENAI_ENDPOINT")
api_version = os.getenv("AZURE_OPENAI_API_VERSION", "2024-05-01-preview")
deployment_name = os.getenv("AZURE_OPENAI_DEPLOYMENT_NAME", "gpt-4o")


def translate_text(text):
    openai = AzureOpenAI(
        azure_endpoint=endpoint,
        api_key=api_key,
        api_version="2024-05-01-preview",
    )
    chat_prompt = [
        {
            "role": "system",
            "content": [
                {
                    "type": "text",
                    "text": """\
                        入力された情報はAzureAdvisorリファレンス情報です。この情報を以下のルールで更新してください。

                        # ルール
                        - なるべくそのままの表現で日本語に翻訳する
                        - ですます調で翻訳する
                        - ?を含む文章は、疑問形で表現しないこと。
                        - true/falseは、そのまま「true」「false」とする
                        - 「可能な値」や「選択可能な値」そのものは、翻訳せずにそのままの表現とする
                        - ()などのカッコ文字は、半角()ではなく、全角（）で表現する
                        - (Attribute)は、（属性）と翻訳する
                        - (Required)は、（必須）と翻訳する
                        - (Optional)は、（オプション）と翻訳する
                        """,
                }
            ],
        },
        {
            "role": "user",
            "content": [
                {
                    "type": "text",
                    "text": f"次の英文を日本語に翻訳してください。 {text}",
                }
            ],
        },
    ]
    completion = openai.chat.completions.create(
        model=deployment_name,
        messages=chat_prompt,
        max_tokens=4096,
        temperature=0,
        top_p=0,
        frequency_penalty=0,
        presence_penalty=0,
    )
    completion_json = json.loads(completion.to_json())  # 修正
    translated_text = completion_json["choices"][0]["message"]["content"]  # 修正
    return translated_text


def is_japanese(text):
    """文字列が日本語を含むか判定する"""
    return bool(re.search(r"[ぁ-んァ-ン一-龯]", text))


# Load start row from .env file
start_row = int(os.getenv("START_ROW", 2))  # デフォルトは2行目


def translate_excel_column(sheet, target_col, add_column, force=False):
    # 列の追加を制御
    new_trans_col = target_col + 1 if not add_column else sheet.max_column + 1
    for row in range(start_row, sheet.max_row + 1):  # START_ROWから開始
        source_cell = sheet.cell(row, target_col)
        target_cell = sheet.cell(row, new_trans_col)
        # 翻訳列が空白でない場合はスキップ
        if source_cell.value and (force or not target_cell.value):
            if is_japanese(source_cell.value):  # 日本語の場合はスキップ
                print(f"Skipped (Japanese): {source_cell.value}")
                continue
            translated_text = translate_text(source_cell.value)
            target_cell.value = translated_text
            print(f"Translated: {source_cell.value} -> {translated_text}")


def translate_excel_file(
    file_path, target_col, sheet_name=None, add_column=False, force=False
):
    print(f"Opening file: {file_path}")  # デバッグ用ログ
    workbook = openpyxl.load_workbook(file_path)
    if sheet_name:  # 指定されたシート名のみ処理
        if sheet_name not in workbook.sheetnames:
            print(
                f"Error: Sheet '{sheet_name}' not found in {file_path}. Available sheets: {workbook.sheetnames}"
            )
            sys.exit(1)
        sheets_to_process = [sheet_name]
    else:  # すべてのシートを処理
        sheets_to_process = workbook.sheetnames

    for sheet_name in sheets_to_process:
        print(
            f"----------- Processing sheet {sheet_name} -----------"
        )  # デバッグ用ログ
        sheet = workbook[sheet_name]
        translate_excel_column(sheet, target_col, add_column, force)
    print(f"Saving file: {file_path}")  # デバッグ用ログ
    workbook.save(file_path)


if __name__ == "__main__":
    if len(sys.argv) < 3 or len(sys.argv) > 6:
        print(
            "Usage: python translate.py <dbfile_path> <target_col> [sheet_name] [--add] [--force]"
        )
        sys.exit(1)

    file_path = sys.argv[1]
    target_col = int(sys.argv[2])  # 指定された列番号を取得
    sheet_name = (
        sys.argv[3] if len(sys.argv) >= 4 and not sys.argv[3].startswith("--") else None
    )
    add_column = "--add" in sys.argv
    force = "--force" in sys.argv

    # file_pathが存在しない場合はエラーを出力して終了
    if not os.path.isfile(file_path):
        print(f"Error: {file_path} not found.")
        sys.exit(1)

    print(
        f"Arguments received - File: {file_path}, Column: {target_col}, Sheet: {sheet_name}, Add Column: {add_column}, Force: {force}"
    )  # デバッグ用ログ
    translate_excel_file(file_path, target_col, sheet_name, add_column, force)
